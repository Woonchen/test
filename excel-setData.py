import os
import openpyxl

os.chdir(r"/Users/nick/Desktop")  # Colab 換路徑使用
wb = openpyxl.Workbook()    # 建立空白的 Excel 活頁簿物件
#wb = openpyxl.load_workbook('oxxo.xlsx')    # 開啟現有的 Excel 活頁簿物件

#s1 = wb.create_sheet('Sheet')
s1 = wb['Sheet']

#X*Y蛇行向下
X = 50
Y = 50
n=0
for i in range(0,Y,2):

        for j in range(X):  #左到右
            n=n+1
            pos = "A"+str(n)    #定義儲存格
            s1[pos].value = str(round(0.64*j,3))+";"+str(round(0.64*i,3))   #取小數三位
        
        for j in range(X-1,-1,-1):  #右到左
            n=n+1
            pos = "A"+str(n)
            s1[pos].value = str(round(0.64*j,3))+";"+str(round(0.64*i,3))
wb.save('Niki_monkey.csv')       # 儲存檔案




